import pandas as pd
import os
import glob
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border,Side
import datetime
# from IQA_analyzer.data.data_file import mistakes_types,classification_mistakes_categories


mistakes_types = ['Wrong Classification',
                'Preapproved brand missed',
                'Parent attribute missed',
                'AAL/Document under review missed',
                'Bundle Toy scenario',
                'CSI status missed',
                'Incorrect mismatch/mismatch missed',
                'Incorrect forgery/forgery missed',
                'Elements from checklist missed - Images',
                'Elements from checklist missed - DOC',
                'Elements from checklist missed - both Images and DOC',
                'Blurb edited incorrectly',
                'Wrong reason code (correct decision and blurb)',
                'Wrong blurb picked (correct decision and RC)',
                'Wrong language of communication',
                '3rd or 4th+ communication wrongly proceeded',
                'Special approach missed']

classification_mistakes_categories = ['Accessories',
                                    'Action figures',
                                    'Art & Craft',
                                    'Baby jumpers & walkers',
                                    'Baby Tethers and Soothers',
                                    'Balloons',
                                    'Beach & Bath Toys',
                                    'Board Games',
                                    'Books',
                                    'Building Blocks & Lego',
                                    'Cards',
                                    'Children luminaries',
                                    'Clothing',
                                    'Collectible',
                                    'Costumes & Accessories',
                                    'Decoration & Party Favor',
                                    'Dices',
                                    'Dollhouse Accessories',
                                    'Dolls & Accessories',
                                    'Educational Toys',
                                    'Inflatable Products & Pool toys',
                                    'Interactive Playmats & Robots',
                                    'Jewelry',
                                    'Kitchen Accessories',
                                    'Lawn/Gardening Sets & Playground Equipment',
                                    'Music instruments',
                                    'Piggy Banks',
                                    'Playmat',
                                    'Plush toys',
                                    'Puzzles',
                                    'RC Toys',
                                    'Ride on Toys',
                                    'Scale Models',
                                    'Slings and catapults',
                                    'Sport Equipment',
                                    'Stationary Products',
                                    'Swings',
                                    'Toy Guns & Accessories',
                                    'Toy Tents and Accessories',
                                    'Toy vehicles',
                                    'Vintage products',
                                    'other']


print('Work in progress...Do not close')

mistakes_columns_names = mistakes_types.copy()

mistakes_columns_names.insert(0,'Alias')

classification_mistakes_columns_names = classification_mistakes_categories.copy()

classification_mistakes_columns_names.insert(0,'Alias/Category')

classification_mistakes_columns_names                              

mistakes_column_title = '[Auditor]Mistake Type'

classification_mistakes_column_title = '[Auditor]Product type - only for classification mistakes'

class Date_Time:
    def __init__(self):
        self.today = self.determine_today()
        self.formatted_today = self.format_today()

    def determine_today(self):
        return datetime.date.today()
    
    def format_today(self):
        return self.today.strftime('%m/%d/%Y')
    

class New_excel_editor:

    def __init__(self,excel_file):
        self.wb = load_workbook(excel_file)
        self.ws = None

    def create_new_sheet(self,sheet_to_create):
        self.wb.create_sheet(sheet_to_create)

    def delete_sheet(self,sheet_to_delete):
        self.wb.remove(sheet_to_delete)

    def add_image_to_excel(self, image_path, cell):
        img = openpyxl.drawing.image.Image(image_path)
        self.ws.add_image(img,cell)

    def save_excel_file(self,excel_file):
        self.wb.save(excel_file)

    def pick_sheet(self,sheet_name):
        self.ws = self.wb[sheet_name]

    def rename_sheet(self,sheet_to_rename,new_name):
        sheet = self.wb[sheet_to_rename]
        sheet.title = new_name

    @staticmethod
    def define_color(start_color,end_color,fill_type):
        fill = PatternFill(start_color=start_color, end_color=end_color,fill_type=fill_type)
        return fill
    
    @staticmethod
    def define_border_style(style):
        border = Border(left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style))
        return border

    def apply_border_style(self,style):

        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = style


    def highlight_above_zero(self,fill_color):

        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int,float)) and cell.value > 0:
                    cell.fill = fill_color

    def highlight_not_empty_column(self,fill_color,column):
        for cell in self.ws[column]:
            if cell.value:
                cell.fill = fill_color

    def highlight_not_empty_row(self,fill_color,row):
        for cell in self.ws[row]:
            if cell.value:
                cell.fill = fill_color
        
class Graphs:

    def __init__(self,data_frame):
        self.data_frame = data_frame
        self.selected_data = None

    def transpose_df(self):
        self.data_frame = self.data_frame.transpose()
    

    def select_data(self,rows,columns):

        if isinstance(self.data_frame,pd.DataFrame):
            self.selected_data = self.data_frame.loc[rows,columns]
        else:
            raise ValueError("Data must be a pandas DataFrame")


    def create_graph(self,columns,title):

        plt.figure(figsize=(8,6))
        plt.bar(columns,self.selected_data.values.flatten(), width=0.5)
        plt.xlabel('Columns')
        plt.ylabel('Values')
        plt.title(title)
        plt.xticks(rotation=90)
        # plt.legend()
        plt.grid(True)
        plt.tight_layout()

    def save_graph(self,name):
        plt.savefig(name)

    def show_graph(self):
        self.show = plt.show()
    

class Counters:

    def __init__(self):
        self.row_counter_1 = 0
        self.row_counter_2 = 0
        self.row_counter_3 = 0
        self.row_counter_4 = 0

    @property
    def _row_counter_1(self):
        return self.row_counter_1
    
    @property
    def _row_counter_2(self):
        return self.row_counter_2
    
    @property
    def _row_counter_3(self):
        return self.row_counter_2
    
class Reviewers:

    def __init__(self,files_names):
        self.files_names = files_names
        self.reviewers_cleaned = self.clean_paths()
        self.reviewers_amount = self.count_reviewers()
        self.first_reviewer = self.reviewers_cleaned_first()
        self.last_reviewer = self.reviewers_cleaned_last()

    
    def clean_paths(self):
        names = [os.path.splitext(os.path.basename(path))[0] for path in self.files_names]
        return names
    
    @property
    def _reviewers_cleaned(self):
        return self.reviewers_cleaned
    
    def reviewers_cleaned_last(self):
        return self.reviewers_cleaned[-1]
    
    def reviewers_cleaned_first(self):
        return self.reviewers_cleaned[0]
    
    def count_reviewers(self):
        count = len(self.reviewers_cleaned)
        return count

    
class Files:

    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.files = self.read_xlsx_files()
        self.number_of_files = self.count_files()

        self.all_reviewers = []

    def clean_files_names(self):
        for file in self.files:
            reviewer = os.path.splitext(file)[0]
            reviewer = reviewer.replace('C:\\New Folder\\','')
            self.add_to_all_reviewers

    def add_to_all_reviewers(self,reviewer):
            self.all_reviewers.append(reviewer)

    def read_xlsx_files(self):
        files = glob.glob(os.path.join(self.folder_path,'*.xlsx'))
        return files
    
    def count_files(self):
        count = len(self.files)
        return count
        
class DataFrame:

    def __init__(self,columns,rows):
        self.columns = columns
        self.rows = rows +1
        self.data = []
        self.data_frame = self.create_data_frame()

    @property
    def _rows(self):
        return self.rows


    def create_data_frame(self):

        data_frame = pd.DataFrame(columns=self.columns)
        for _ in range(int(self.rows)):
            data_frame = pd.concat([data_frame, pd.DataFrame(columns=self.columns)], ignore_index=True)
            return data_frame

    def add_to_data_frame(self,value,column,row):
        self.data_frame.at[row,column] = value

    def sum_in_data_frame(self,columns,rows):

        if isinstance(rows,int):
            selected_rows = [rows]

        elif isinstance(rows,tuple) and len(rows) == 2:
            selected_rows = range(rows[0], rows[1] + 1)
        
        else:
            raise ValueError('Invalid input for rows')
        
        if isinstance(columns,str):
            selected_columns = [columns]

        elif isinstance(columns, list):
            selected_columns = columns
        
        else:
            raise ValueError('Invalid input for columns')
        
        selected_values = self.data_frame.loc[selected_rows,selected_columns]

        return selected_values.sum().sum()
    
    def pick_data(self,columns,rows):

        if isinstance(rows,int):
            selected_rows = [rows]

        elif isinstance(rows,tuple) and len(rows) == 2:
            selected_rows = range(rows[0], rows[1] + 1)
        
        else:
            raise ValueError('Invalid input for rows')
        
        if isinstance(columns,str):
            selected_columns = [columns]

        elif isinstance(columns, list):
            selected_columns = columns
        
        else:
            raise ValueError('Invalid input for columns')
        
        selected_values = self.data_frame.loc[selected_rows,selected_columns]

        return selected_values
        

    def save_to_excel(self,file_name):
        self.data_frame.to_excel(file_name, index=False)

    def save_to_existing_excel(self,file_name,sheet_name):
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            self.data_frame.to_excel(writer, sheet_name=sheet_name, index=False)


class ExcelDataAnalyzer:

    def __init__(self,file_path):
        self.df = pd.read_excel(file_path)
        self.file_path = file_path
        self.reviewer = self.set_reviewer

    def find_value(self,row,column):
        try:
            print(value)
            value = self.df.at[row, column]
            return value
        except KeyError:
            print(f"Value not found. Row'{row}' or Column '{column}' does not exist")
            return None
        
    @property
    def set_reviewer(self):
        reviewer = os.path.splitext(self.file_path)[0]
        reviewer = reviewer.replace('C:\\New Folder\\','')
        return reviewer
    
class Column(ExcelDataAnalyzer):

    def __init__(self,column,file_path):
        self.column = column
        super().__init__(file_path)
        
    def count_occurance(self, search_value):
        return(self.df[self.column] == search_value).sum()


if __name__ == "__main__":

    counters = Counters()

    IQA_files = Files('C:\\New Folder')

    reviewers = Reviewers(IQA_files.files)

    DR_mistakes_DF = DataFrame(columns = mistakes_columns_names, rows = IQA_files.number_of_files)

    Classification_mistakes_DF = DataFrame(columns = [], rows = len(classification_mistakes_categories) )

    for trainee in IQA_files.files:

        excel_analyzer = ExcelDataAnalyzer(trainee)
        
        DR_mistakes_DF.add_to_data_frame(excel_analyzer.reviewer,'Alias', counters.row_counter_1)
        
        Mistakes_column = Column(mistakes_column_title,trainee)
    
        for mistake in mistakes_types:

            mistakes_amount = Mistakes_column.count_occurance(mistake)

            DR_mistakes_DF.add_to_data_frame(mistakes_amount,mistake,counters.row_counter_1)
        
        
        counters.row_counter_1 += 1

        DR_mistakes_DF.add_to_data_frame('SUM','Alias',counters.row_counter_1)

        Toy_categories_column = Column(classification_mistakes_column_title,trainee)


        for toy_category in classification_mistakes_categories:

            Classification_mistakes_DF.add_to_data_frame(toy_category,'Alias/Category', counters.row_counter_2)

            count = Toy_categories_column.count_occurance(toy_category)

            Classification_mistakes_DF.add_to_data_frame(count, excel_analyzer.reviewer, counters.row_counter_2)

            counters.row_counter_2 += 1
        
        counters.row_counter_2 = 0


    for mistake in mistakes_types:

        column = mistake
        rows = (0,counters.row_counter_1)
        sum = DR_mistakes_DF.sum_in_data_frame(column,rows)

        DR_mistakes_DF.add_to_data_frame(sum,mistake,counters.row_counter_1)

    for toy_category in classification_mistakes_categories:

        sum = Classification_mistakes_DF.sum_in_data_frame(reviewers.reviewers_cleaned,counters.row_counter_4)


        Classification_mistakes_DF.add_to_data_frame(sum,'Sum',int(counters.row_counter_4))

        counters.row_counter_4 += 1

Date_time = Date_Time()

final_excel_name = f'IQA_results_{Date_time.formatted_today}.xlsx'

final_excel_name = final_excel_name.replace('/','.')

DR_mistakes_DF.save_to_excel(final_excel_name)

Classification_mistakes_DF.save_to_existing_excel(final_excel_name,'classification_mistakes')

DR_mistakes_graph = Graphs(DR_mistakes_DF.data_frame)

DR_mistakes_graph.select_data([counters.row_counter_1],mistakes_types)

DR_mistakes_graph.create_graph(mistakes_types,'DR mistakes types')

Excel_editor=New_excel_editor(final_excel_name)

Excel_editor.create_new_sheet('DR_mistakes_graph')

Excel_editor.create_new_sheet('Classification_mistakes_graph')

Excel_editor.rename_sheet('Sheet1','DR_mistakes')

DR_mistakes_graph.save_graph('DR_mistakes_graph')

Excel_editor.pick_sheet('DR_mistakes_graph')

Excel_editor.highlight_above_zero

Excel_editor.add_image_to_excel('DR_mistakes_graph.png','A1')

Classification_graph = Graphs(Classification_mistakes_DF.data_frame)

Classification_graph.transpose_df()

Classification_graph.select_data('Sum',range(0,42))

Classification_graph.create_graph(classification_mistakes_categories,'Classification mistakes types')

Classification_graph.save_graph('Classification_mistakes.png')

Excel_editor.pick_sheet('Classification_mistakes_graph')

Excel_editor.add_image_to_excel('Classification_mistakes.png', 'A1')

Excel_editor.pick_sheet('DR_mistakes')

red = Excel_editor.define_color('FFCCCB','FFCCCB','solid')

orange = Excel_editor.define_color('FFD966','FFD966','solid')

blue = Excel_editor.define_color('ADD8E6','ADD8E6','solid')

thick = Excel_editor.define_border_style('thick')

Excel_editor.highlight_above_zero(red)

Excel_editor.highlight_not_empty_column(blue,'A')

Excel_editor.highlight_not_empty_row(orange,1)

Excel_editor.apply_border_style(thick)

Excel_editor.pick_sheet('classification_mistakes')

Excel_editor.highlight_not_empty_column(blue,'A')

Excel_editor.highlight_not_empty_row(orange,1)

Excel_editor.highlight_above_zero(red)

Excel_editor.apply_border_style(thick)

Excel_editor.save_excel_file(final_excel_name)

print('Success! Results file was created next to this programm location!')



